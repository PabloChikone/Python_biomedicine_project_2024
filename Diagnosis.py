import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment, Side

# Считывание из файла
def read_csv(file_name):
    data = pd.read_csv(file_name, sep=';')
    return data
def read_excel(file_name):
    data = pd.read_excel(file_name)
    return data

# Запись данных в файл
def write_csv(data, file_name):
    data.to_csv(file_name, index=False, sep=';')
def write_excel(data, file_name):
    data.to_excel(file_name, index=False)

# функция по формированию уникального списка значений по выбранному столбцу из файла csv
def unique_values_column(data):
    unique_values_data = pd.DataFrame()  # Создаем пустой DataFrame для уникальных значений

    # Проходим по каждому столбцу в исходном DataFrame
    for column in data.columns:
        # Получаем уникальные значения для текущего столбца и создаем DataFrame
        temp_df = pd.DataFrame(data[column].unique(), columns=[column])

        # Добавляем временный DataFrame к DataFrame с уникальными значениями
        unique_values_data = pd.concat([unique_values_data, temp_df], ignore_index=True)

    return unique_values_data

# функция по формированию подсчета значений по выбранному столбцу из файла
def count_and_merge_data(dataframe_file_unique, column_name_unique, dataframe_file_current, column_name_current):
    # Создаем новый датафрейм для результатов
    merged_dataframe = pd.DataFrame()

    # Копируем столбец 'diagnosis_mkb_code4' из первого датафрейма в новый датафрейм
    merged_dataframe[column_name_unique] = dataframe_file_unique[column_name_unique]

    # Создаем второй столбец 'diagnosis_mkb_code4_merge' и подсчитываем количество повторений
    # Используем метод merge для объединения датафреймов по столбцу 'diagnosis_mkb_code4'
    merged_dataframe[column_name_unique + '_merge'] = dataframe_file_current.groupby(column_name_current).size().reset_index(name='count')['count']

    return merged_dataframe

# функция расширения данных по группировке диагнозов
def data_unique_column_append(dataframe_file_current, dataframe_file_mkb, column_name):
    # Создаем копию dataframe_file_current
    append_dataframe = dataframe_file_current.copy()

    # Создаем пустой столбец 'group_disease' в новом датафрейме
    append_dataframe['group_disease'] = ""

    # Функция для поиска совпадения в dataframe_file_mkb
    def find_group_name(value):
        # Проходим по каждому столбцу в dataframe_file_mkb
        for col in dataframe_file_mkb.columns:
            # Проверяем наличие значения в текущем столбце
            if value in dataframe_file_mkb[col].values:
                # Возвращаем имя столбца, если найдено совпадение
                return col
        # Если совпадение не найдено, возвращаем None
        return None

    # Применяем функцию find_group_name к столбцу column_name и сохраняем результат в 'group_disease'
    append_dataframe['group_disease'] = append_dataframe[column_name].apply(find_group_name)

    return append_dataframe

# функция расширения данных с указанием возраста пациента
def data_unique_column_append_age(dataframe_file_current, column_name):
    # Создаем копию dataframe_file_current
    append_dataframe = dataframe_file_current.copy()

    # Заполняем пустые значения в столбце 'case_patient_birthdate' значением 'NaT'
    append_dataframe[column_name] = append_dataframe[column_name].fillna(pd.NaT)

    # Преобразуем столбец 'case_patient_birthdate' в формат даты
    append_dataframe[column_name] = pd.to_datetime(append_dataframe[column_name], errors='coerce', format='%d.%m.%Y')

    # Рассчитываем возраст пациентов к 31.12.2024, игнорируя пустые значения
    append_dataframe['age_pacient'] = (datetime(2024, 12, 31) - append_dataframe[column_name]).dt.days // 365

    return append_dataframe


# функция рассчета по data_append количество заболеваний по фильтрам
def filter_data_append(dataframe_filter, data_append):
    # Создаем новый столбец 'Количество случаев заболевания в 2019' с нулевыми значениями
    columns_to_fill_with_zeros = [
        'Количество случаев заболевания в 2019',
        'Количество случаев заболевания в 2020',
        'Количество случаев заболевания в 2021',
        'Количество случаев заболевания в 2022',
        'Количество случаев заболевания в 2023',
        'Количество случаев заболевания в 1 квартал 2024',
        'Случаев заболевания мужчинами',
        'Случаев заболевания женщинами',
        'Количество случаев заболевания среди пациентов до 20 лет',
        'Количество случаев заболевания среди пациентов до 30 лет',
        'Количество случаев заболевания среди пациентов до 40 лет',
        'Количество случаев заболевания среди пациентов до 50 лет',
        'Количество случаев заболевания среди пациентов до 60 лет',
        'Количество случаев заболевания среди пациентов старше 60 лет'
    ]
    # Заполняем указанные столбцы нулями
    dataframe_filter[columns_to_fill_with_zeros] = 0

    # Преобразуем столбцы с датой в datetime
    data_append['case_fiemk_date'] = pd.to_datetime(data_append['case_fiemk_date'], format='%d.%m.%Y')
    data_append['case_patient_birthdate'] = pd.to_datetime(data_append['case_patient_birthdate'], format='%d.%m.%Y')

    # Для каждой строки в dataframe_filter
    for index, row in dataframe_filter.iterrows():
        # Подсчитываем количество соответствий в data_append по дате установления диагноза
        count_2019 = (data_append[data_append['case_fiemk_date'].dt.year == 2019]["group_disease"].value_counts().get(row["Наименование группы болезни"], 0))
        count_2020 = (data_append[data_append['case_fiemk_date'].dt.year == 2020]["group_disease"].value_counts().get(row["Наименование группы болезни"], 0))
        count_2021 = (data_append[data_append['case_fiemk_date'].dt.year == 2021]["group_disease"].value_counts().get(row["Наименование группы болезни"], 0))
        count_2022 = (data_append[data_append['case_fiemk_date'].dt.year == 2022]["group_disease"].value_counts().get(row["Наименование группы болезни"], 0))
        count_2023 = (data_append[data_append['case_fiemk_date'].dt.year == 2023]["group_disease"].value_counts().get(row["Наименование группы болезни"], 0))
        count_2024 = (data_append[data_append['case_fiemk_date'].dt.year == 2024]["group_disease"].value_counts().get(row["Наименование группы болезни"], 0))
        # Подсчитываем количество соответствий в data_append по полу
        count_men = (data_append[data_append['case_patient_gender_ru'] == "мужской"]["group_disease"].value_counts().get(row["Наименование группы болезни"], 0))
        count_women = (data_append[data_append['case_patient_gender_ru'] == "женский"]["group_disease"].value_counts().get(row["Наименование группы болезни"], 0))
        # Подсчитываем количество соответствий в data_append по дате рождения
        count_dr_before_20 = (data_append[(data_append['age_pacient'] >= 0) & (data_append['age_pacient'] < 20)]["group_disease"].value_counts().get(row["Наименование группы болезни"], 0))
        count_dr_before_30 = (data_append[(data_append['age_pacient'] >= 20) & (data_append['age_pacient'] < 30)]["group_disease"].value_counts().get(row["Наименование группы болезни"], 0))
        count_dr_before_40 = (data_append[(data_append['age_pacient'] >= 30) & (data_append['age_pacient'] < 40)]["group_disease"].value_counts().get(row["Наименование группы болезни"], 0))
        count_dr_before_50 = (data_append[(data_append['age_pacient'] >= 40) & (data_append['age_pacient'] < 50)]["group_disease"].value_counts().get(row["Наименование группы болезни"], 0))
        count_dr_before_60 = (data_append[(data_append['age_pacient'] >= 50) & (data_append['age_pacient'] < 60)]["group_disease"].value_counts().get(row["Наименование группы болезни"], 0))
        count_dr_after_60  = (data_append[(data_append['age_pacient'] >= 60)]["group_disease"].value_counts().get(row["Наименование группы болезни"], 0))

        # Обновляем значения
        dataframe_filter.at[index, "Количество случаев заболевания в 2019"] = count_2019
        dataframe_filter.at[index, "Количество случаев заболевания в 2020"] = count_2020
        dataframe_filter.at[index, "Количество случаев заболевания в 2021"] = count_2021
        dataframe_filter.at[index, "Количество случаев заболевания в 2022"] = count_2022
        dataframe_filter.at[index, "Количество случаев заболевания в 2023"] = count_2023
        dataframe_filter.at[index, "Количество случаев заболевания в 1 квартал 2024"] = count_2024

        dataframe_filter.at[index, "Случаев заболевания мужчинами"] = count_men
        dataframe_filter.at[index, "Случаев заболевания женщинами"] = count_women

        dataframe_filter.at[index, "Количество случаев заболевания среди пациентов до 20 лет"] = count_dr_before_20
        dataframe_filter.at[index, "Количество случаев заболевания среди пациентов до 30 лет"] = count_dr_before_30
        dataframe_filter.at[index, "Количество случаев заболевания среди пациентов до 40 лет"] = count_dr_before_40
        dataframe_filter.at[index, "Количество случаев заболевания среди пациентов до 50 лет"] = count_dr_before_50
        dataframe_filter.at[index, "Количество случаев заболевания среди пациентов до 60 лет"] = count_dr_before_60
        dataframe_filter.at[index, "Количество случаев заболевания среди пациентов старше 60 лет"] = count_dr_after_60



    return dataframe_filter



# функция форматирования файла Excel для красоты
def format_excel_file(input_file_path):
    # Загрузка файла Excel
    wb = load_workbook(input_file_path)
    ws = wb.active

    # Создание стилей для форматирования
    font = Font(name='Times New Roman', size=10)
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Применение форматирования к каждой ячейке с данными
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.font = font
                cell.border = border
                cell.alignment = alignment

    # Установка ширины столбцов
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        ws.column_dimensions[column].width = 25

    # Автоподбор высоты строки
    for row in ws.rows:
        for cell in row:
            if cell.value is not None:
                cell.alignment = alignment
                ws.row_dimensions[cell.row].auto_size = True

    # Сохранение изменений в файл
    wb.save(input_file_path)






















